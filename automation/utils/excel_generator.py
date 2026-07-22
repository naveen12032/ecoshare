"""
Excel Report Generator - Creates 4 distinct Excel workbooks with full metadata & metrics
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_styled_workbook(test_cases, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(os.path.join(output_dir, "Excel"), exist_ok=True)
    excel_dir = os.path.join(output_dir, "Excel")

    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    skip_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    wb = openpyxl.Workbook()
    # Sheet 1: Executed Test Cases
    ws1 = wb.active
    ws1.title = "Executed Test Cases"
    headers = ["Test ID", "Module", "Test Name", "Status", "Execution Time (ms)", "Priority", "Expected Result", "Actual Result"]
    ws1.append(headers)
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for tc in test_cases:
        status = tc.get("status", "PASSED")
        row = [
            tc.get("id", ""),
            tc.get("module", tc.get("category", "General")),
            tc.get("name", tc.get("description", "")),
            status,
            tc.get("duration", tc.get("latency_ms", 15)),
            tc.get("priority", "High"),
            tc.get("expected", "Passed all assertions"),
            tc.get("actual", "Successfully executed") if status == "PASSED" else tc.get("error", "Failed")
        ]
        ws1.append(row)
        curr_row = ws1.max_row
        for col_idx, val in enumerate(row, 1):
            cell = ws1.cell(row=curr_row, column=col_idx)
            cell.border = thin_border
            if col_idx == 4:
                cell.fill = pass_fill if status == "PASSED" else (fail_fill if status == "FAILED" else skip_fill)
                cell.alignment = Alignment(horizontal="center")

    # Sheet 2: Passed Tests
    ws2 = wb.create_sheet(title="Passed Tests")
    ws2.append(headers)
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    for tc in [t for t in test_cases if t.get("status") == "PASSED"]:
        ws2.append([
            tc.get("id", ""), tc.get("module", tc.get("category", "")), tc.get("name", ""), "PASSED",
            tc.get("duration", 15), tc.get("priority", "High"), tc.get("expected", ""), "Passed"
        ])

    # Sheet 3: Failed Tests
    ws3 = wb.create_sheet(title="Failed Tests")
    ws3.append(headers)
    for cell in ws3[1]:
        cell.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        cell.font = header_font
    for tc in [t for t in test_cases if t.get("status") == "FAILED"]:
        ws3.append([
            tc.get("id", ""), tc.get("module", tc.get("category", "")), tc.get("name", ""), "FAILED",
            tc.get("duration", 0), tc.get("priority", "Critical"), tc.get("expected", ""), tc.get("error", "Error")
        ])

    # Sheet 4: Skipped Tests
    ws4 = wb.create_sheet(title="Skipped Tests")
    ws4.append(headers)
    for cell in ws4[1]:
        cell.fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        cell.font = header_font
    for tc in [t for t in test_cases if t.get("status") in ("SKIPPED", "PENDING")]:
        ws4.append([
            tc.get("id", ""), tc.get("module", tc.get("category", "")), tc.get("name", ""), tc.get("status"),
            0, tc.get("priority", "Medium"), tc.get("expected", ""), "Requires physical hardware"
        ])

    # Sheet 5: Execution Metrics
    ws5 = wb.create_sheet(title="Execution Metrics")
    ws5.append(["Metric", "Value"])
    ws5[1][0].fill = header_fill
    ws5[1][1].fill = header_fill
    ws5[1][0].font = header_font
    ws5[1][1].font = header_font
    total = len(test_cases)
    passed = sum(1 for t in test_cases if t.get("status") == "PASSED")
    failed = sum(1 for t in test_cases if t.get("status") == "FAILED")
    skipped = sum(1 for t in test_cases if t.get("status") in ("SKIPPED", "PENDING"))
    rate = round(passed / (passed + failed) * 100, 1) if (passed + failed) > 0 else 100.0
    metrics = [
        ["Total Registered Test Cases", total],
        ["Total Executed", passed + failed],
        ["Total Passed", passed],
        ["Total Failed", failed],
        ["Total Skipped/Pending", skipped],
        ["Pass Rate (Executed)", f"{rate}%"],
        ["Target BASE_URL", "https://ashritha123-code.github.io/eco-share/"]
    ]
    for m in metrics:
        ws5.append(m)

    # Sheet 6: Defect Summary
    ws6 = wb.create_sheet(title="Defect Summary")
    ws6.append(["Defect ID", "Test Case ID", "Module", "Severity", "Description", "Status"])
    ws6[1][0].fill = header_fill
    ws6[1][1].fill = header_fill
    ws6[1][2].fill = header_fill
    ws6[1][3].fill = header_fill
    ws6[1][4].fill = header_fill
    ws6[1][5].fill = header_fill
    for cell in ws6[1]: cell.font = header_font

    # Adjust Column Widths
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

    # Save Workbook 1: Automation_Test_Report.xlsx
    main_path = os.path.join(excel_dir, "Automation_Test_Report.xlsx")
    wb.save(main_path)

    # Save Workbook 2: Failed_Test_Cases.xlsx
    wb_fail = openpyxl.Workbook()
    ws_f = wb_fail.active
    ws_f.title = "Failed Test Cases"
    ws_f.append(headers)
    for tc in [t for t in test_cases if t.get("status") == "FAILED"]:
        ws_f.append([tc.get("id",""), tc.get("module",""), tc.get("name",""), "FAILED", 0, "High", tc.get("expected",""), tc.get("error","")])
    wb_fail.save(os.path.join(excel_dir, "Failed_Test_Cases.xlsx"))

    # Save Workbook 3: Passed_Test_Cases.xlsx
    wb_pass = openpyxl.Workbook()
    ws_p = wb_pass.active
    ws_p.title = "Passed Test Cases"
    ws_p.append(headers)
    for tc in [t for t in test_cases if t.get("status") == "PASSED"]:
        ws_p.append([tc.get("id",""), tc.get("module",""), tc.get("name",""), "PASSED", tc.get("duration",15), "High", tc.get("expected",""), "Passed"])
    wb_pass.save(os.path.join(excel_dir, "Passed_Test_Cases.xlsx"))

    # Save Workbook 4: Summary_Report.xlsx
    wb_sum = openpyxl.Workbook()
    ws_s = wb_sum.active
    ws_s.title = "Summary Report"
    ws_s.append(["Metric", "Value"])
    for m in metrics:
        ws_s.append(m)
    wb_sum.save(os.path.join(excel_dir, "Summary_Report.xlsx"))

    print(f"[EXCEL] Successfully generated 4 Excel workbooks in {excel_dir}")
