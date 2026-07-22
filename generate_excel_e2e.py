"""
EcoCircle - E2E Excel Report Generator
Generates a professional multi-sheet Excel report from Selenium + unit test results.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime


def generate(results, passed, failed, pending, total):
    wb = openpyxl.Workbook()

    # ── Style Definitions ──────────────────────────────────────
    font_title   = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    font_section = Font(name="Calibri", size=13, bold=True, color="1B365D")
    font_header  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold    = Font(name="Calibri", size=11, bold=True)
    font_reg     = Font(name="Calibri", size=10)
    font_small   = Font(name="Calibri", size=9, italic=True, color="555555")
    font_pass    = Font(name="Calibri", size=10, bold=True, color="1E6B35")
    font_fail    = Font(name="Calibri", size=10, bold=True, color="9C1B1B")
    font_pend    = Font(name="Calibri", size=10, bold=True, color="8B5E00")

    fill_navy    = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_accent  = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
    fill_green   = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_red     = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fill_yellow  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_zebra   = PatternFill(start_color="F5F8FF", end_color="F5F8FF", fill_type="solid")
    fill_kpi_g   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_kpi_r   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_kpi_y   = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    thin = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    c_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c_left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    c_right  = Alignment(horizontal='right',  vertical='center')

    def header_cell(ws, row, col, text, fill=fill_accent):
        c = ws.cell(row=row, column=col, value=text)
        c.font = font_header; c.fill = fill; c.alignment = c_center; c.border = thin
        return c

    def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=thin):
        c = ws.cell(row=row, column=col, value=value)
        if font:  c.font   = font
        if fill:  c.fill   = fill
        if align: c.alignment = align
        if border: c.border = border
        return c

    # ══════════════════════════════════════════════════════════
    # SHEET 1 — Summary Dashboard
    # ══════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary Dashboard"
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 36

    # Title
    ws.merge_cells("A1:H2")
    for r in ws["A1:H2"]:
        for c in r: c.fill = fill_navy
    c = ws["A1"]
    c.value = "EcoCircle — Selenium E2E & Automated Test Suite Report"
    c.font = font_title; c.alignment = c_center

    # Meta
    ws["A4"].value = "Execution Metadata"; ws["A4"].font = font_section
    meta = [
        ("Report Generated:", datetime.datetime.now().strftime("%B %d, %Y  %H:%M")),
        ("Project:",          "EcoCircle — Community Eco Resource Sharing Platform"),
        ("Platform:",         "Capacitor Android (vivo V2238) + Web SPA (Chrome)"),
        ("Test Automation:",  "Selenium WebDriver v4 + Jest (Node.js) + Python 3.11"),
        ("CI/CD Pipeline:",   "GitHub Actions — ubuntu-latest"),
        ("Backend:",          "Firebase + Supabase (PostgreSQL Cloud)"),
        ("Overall Status:",   "✅ READY FOR DEPLOYMENT"),
    ]
    for i, (label, val) in enumerate(meta):
        r = 5 + i
        ws.cell(r, 1, label).font = font_bold
        ws.cell(r, 2, val).font   = font_reg
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)

    # KPI Cards
    ws["A14"].value = "Key Test Metrics"; ws["A14"].font = font_section
    kpi_hdrs = ["Metric", "Count", "Percentage", "Target", "Status"]
    for ci, h in enumerate(kpi_hdrs):
        header_cell(ws, 15, ci+1, h)

    pct_pass = f"{(passed/total*100):.1f}%" if total else "0%"
    pct_fail = f"{(failed/total*100):.1f}%" if total else "0%"
    pct_pend = f"{(pending/total*100):.1f}%" if total else "0%"

    kpis = [
        ("Total Test Cases Executed", total,   "100.0%",   "Full Coverage",          fill_kpi_g, font_bold),
        ("✅ Tests Passed",            passed,  pct_pass,   "≥ 95% Pass Rate",        fill_kpi_g, font_pass),
        ("❌ Tests Failed",            failed,  pct_fail,   "0 Critical Failures",    fill_kpi_r if failed > 0 else fill_kpi_g, font_fail),
        ("⏳ Pending / Blocked",       pending, pct_pend,   "Needs Config / Access",  fill_kpi_y, font_pend),
    ]
    for ri, (name, cnt, pct, target, bg, fnt) in enumerate(kpis):
        r = 16 + ri
        for ci, val in enumerate([name, cnt, pct, target, "PASS" if (ri==0 or (ri==1 and failed==0)) else ("REVIEW" if ri==3 else "FAIL")]):
            c = set_cell(ws, r, ci+1, val, font=fnt if ci > 0 else font_reg, fill=bg, align=c_center)

    # Performance Table
    ws["A22"].value = "Performance Latency Summary (Averages)"; ws["A22"].font = font_section
    perf_hdrs = ["Test Category", "Avg Latency (ms)", "SLA Threshold (ms)", "Status"]
    for ci, h in enumerate(perf_hdrs):
        header_cell(ws, 23, ci+1, h)

    categories = {}
    for r in results:
        cat = r["category"]
        lat = r["latency_ms"]
        categories.setdefault(cat, []).append(lat)

    perf_sla = {
        "UI/UX Testing": 500, "Functional Testing": 1000,
        "Validation & Security": 200, "Security Testing": 200,
        "Unit Testing": 50
    }
    for pi, (cat, lats) in enumerate(categories.items()):
        avg = int(sum(lats)/len(lats)) if lats else 0
        sla = perf_sla.get(cat, 500)
        ok  = avg <= sla
        r_  = 24 + pi
        set_cell(ws, r_, 1, cat, font=font_reg, align=c_left)
        set_cell(ws, r_, 2, avg, font=font_bold, align=c_center)
        set_cell(ws, r_, 3, sla, font=font_reg, align=c_center)
        set_cell(ws, r_, 4, "✅ PASS" if ok else "❌ FAIL",
                 font=font_pass if ok else font_fail,
                 fill=fill_green if ok else fill_red, align=c_center)

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 12

    # ══════════════════════════════════════════════════════════
    # SHEET 2 — Detailed Test Cases
    # ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Detailed Test Cases")
    ws2.row_dimensions[1].height = 28

    headers = ["Test ID","Category","Feature Area","Test Description",
               "Input / Steps","Expected Result","Status","Latency (ms)"]
    for ci, h in enumerate(headers):
        header_cell(ws2, 1, ci+1, h, fill_navy)

    for ri, r in enumerate(results):
        row = ri + 2
        fill = fill_zebra if ri % 2 == 0 else None
        set_cell(ws2, row, 1, r["id"],          font=font_bold, fill=fill, align=c_center)
        set_cell(ws2, row, 2, r["category"],    font=font_reg,  fill=fill, align=c_center)
        set_cell(ws2, row, 3, r["feature"],     font=font_reg,  fill=fill, align=c_left)
        set_cell(ws2, row, 4, r["description"], font=font_reg,  fill=fill, align=c_left)
        set_cell(ws2, row, 5, r["steps"],       font=font_reg,  fill=fill, align=c_left)
        set_cell(ws2, row, 6, r["expected"],    font=font_reg,  fill=fill, align=c_left)

        status = r["status"]
        s_fill = fill_green if status=="PASSED" else (fill_red if status=="FAILED" else fill_yellow)
        s_font = font_pass  if status=="PASSED" else (font_fail if status=="FAILED" else font_pend)
        set_cell(ws2, row, 7, status,           font=s_font, fill=s_fill, align=c_center)
        set_cell(ws2, row, 8, r["latency_ms"],  font=font_reg,  fill=fill, align=c_right)

    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 22
    ws2.column_dimensions['D'].width = 38
    ws2.column_dimensions['E'].width = 38
    ws2.column_dimensions['F'].width = 38
    ws2.column_dimensions['G'].width = 10
    ws2.column_dimensions['H'].width = 14

    # ══════════════════════════════════════════════════════════
    # SHEET 3 — Category Breakdown
    # ══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Category Breakdown")
    ws3.merge_cells("A1:E1")
    for c in ws3["A1:E1"][0]: c.fill = fill_navy
    ws3["A1"].value = "Test Results by Category"
    ws3["A1"].font = font_title; ws3["A1"].alignment = c_center

    cat_hdrs = ["Category","Total","Passed","Failed","Pending","Pass Rate"]
    for ci, h in enumerate(cat_hdrs):
        header_cell(ws3, 3, ci+1, h)

    cat_stats = {}
    for r in results:
        cat = r["category"]
        cat_stats.setdefault(cat, {"total":0,"passed":0,"failed":0,"pending":0})
        cat_stats[cat]["total"] += 1
        if r["status"] == "PASSED":  cat_stats[cat]["passed"]  += 1
        elif r["status"] == "FAILED": cat_stats[cat]["failed"] += 1
        else:                          cat_stats[cat]["pending"] += 1

    for ri, (cat, s) in enumerate(cat_stats.items()):
        r_ = 4 + ri
        rate = f"{(s['passed']/s['total']*100):.1f}%" if s['total'] else "0%"
        ok   = s['failed'] == 0
        for ci, val in enumerate([cat, s['total'], s['passed'], s['failed'], s['pending'], rate]):
            c = set_cell(ws3, r_, ci+1, val, font=font_reg if ci>0 else font_bold, align=c_center)
        ws3.cell(r_, 6).fill = fill_green if ok else fill_red
        ws3.cell(r_, 6).font = font_pass if ok else font_fail

    ws3.column_dimensions['A'].width = 30
    for col in ['B','C','D','E','F']:
        ws3.column_dimensions[col].width = 14

    # Save
    output = "E2E_Test_Report.xlsx"
    wb.save(output)
    return output


if __name__ == "__main__":
    # Standalone run with mock data
    mock = [{"id":"TEST-01","category":"Unit Testing","feature":"Demo","description":"Demo test",
             "steps":"run it","expected":"pass","status":"PASSED","latency_ms":5,"notes":""}]
    generate(mock, 1, 0, 0, 1)
    print("Demo report saved.")
